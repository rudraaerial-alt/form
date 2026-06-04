import os
from datetime import datetime, timezone
from pathlib import Path

from flask import Flask, jsonify, render_template, request
from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_FILE_PATH = DATA_DIR / "pe_form_data.xlsx"

HEADERS = [
    "S.No",
    "PE / Non-PE",
    "Category",
    "Executor",
    "Total Node Count",
    "Site / Cell",
    "Circle",
    "Date of PE",
    "PE Purpose",
    "PE Purpose - Sub2",
    "Activity Initiator",
    "Activity Owner",
    "Activity Executor",
    "Deploy Approval",
    "NPO KPI Approval",
    "NPO AT Approval",
    "TSS Approval",
    "Status",
    "Domain",
    "Remarks1",
    "Remarks2",
    "Remarks",
    "Created At (UTC)",
]

app = Flask(__name__, template_folder="templates", static_folder="static")


def ensure_workbook() -> None:
    if not EXCEL_FILE_PATH.exists():
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "PE Data"
        worksheet.append(HEADERS)
        workbook.save(EXCEL_FILE_PATH)
        workbook.close()


def list_to_string(value):
    if isinstance(value, list):
        return ", ".join(str(item).strip() for item in value if str(item).strip())
    if value is None:
        return ""
    return str(value).strip()


def get_next_sno() -> int:
    ensure_workbook()
    workbook = load_workbook(EXCEL_FILE_PATH)
    worksheet = workbook.active
    next_sno = max(worksheet.max_row, 1)
    workbook.close()
    return next_sno


def append_to_excel(data: dict) -> None:
    ensure_workbook()
    workbook = load_workbook(EXCEL_FILE_PATH)
    worksheet = workbook.active
    row = [
        data.get("sno", ""),
        list_to_string(data.get("peNonPe")),
        list_to_string(data.get("category")),
        list_to_string(data.get("executor")),
        data.get("totalNodeCount", ""),
        list_to_string(data.get("siteCell")),
        list_to_string(data.get("circle")),
        data.get("dateOfPe", ""),
        list_to_string(data.get("pePurpose")),
        list_to_string(data.get("pePurposeSub2")),
        list_to_string(data.get("activityInitiator")),
        list_to_string(data.get("activityOwner")),
        list_to_string(data.get("activityExecutor")),
        list_to_string(data.get("deployApproval")),
        list_to_string(data.get("npoKpiApproval")),
        list_to_string(data.get("npoAtApproval")),
        list_to_string(data.get("tssApproval")),
        list_to_string(data.get("status")),
        list_to_string(data.get("domain")),
        list_to_string(data.get("remarks1")),
        list_to_string(data.get("remarks2")),
        data.get("remarks", ""),
        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
    ]
    worksheet.append(row)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def validate_payload(data: dict):
    required_single = {
        "peNonPe": "PE / Non-PE",
        "category": "Category",
        "executor": "Executor",
        "siteCell": "Site / Cell",
    }
    if not str(data.get("sno", "")).strip():
        return "S.No missing hai."
    for key, label in required_single.items():
        value = data.get(key, [])
        if isinstance(value, list):
            if not value:
                return f"{label} select karein."
        elif not str(value).strip():
            return f"{label} select karein."
    return None


@app.get("/")
def home():
    ensure_workbook()
    return render_template("Form.html")


@app.get("/health")
def health():
    return jsonify({"status": "ok"})


@app.get("/next-sno")
def next_sno():
    try:
        return jsonify({"next_sno": get_next_sno()})
    except Exception as exc:
        return jsonify({"next_sno": "", "error": str(exc)}), 500


@app.post("/submit")
def submit():
    try:
        data = request.get_json(force=True) or {}
        error_message = validate_payload(data)
        if error_message:
            return jsonify({"success": False, "message": error_message}), 400
        append_to_excel(data)
        return jsonify(
            {
                "success": True,
                "message": "Data Excel me successfully save ho gaya.",
                "file": str(EXCEL_FILE_PATH.name),
            }
        )
    except Exception as exc:
        return jsonify({"success": False, "message": str(exc)}), 500


if __name__ == "__main__":
    ensure_workbook()
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port, debug=False)
