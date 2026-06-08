import os
import traceback
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote

import msal
import requests
from flask import Flask, jsonify, render_template, request
from openpyxl import Workbook, load_workbook

# -----------------------------
# Local storage
# -----------------------------
BASE_DIR = Path("/tmp")
DATA_DIR = BASE_DIR / "data"
EXCEL_FILE_PATH = DATA_DIR / "pe_form_data.xlsx"

HEADERS = [
    "S.No",
    "PE / Non-PE",
    "Category",
    "Executor",
    "Total Node Count",
    "Site / Cell",
    "Circle",
    "Date of PE",
    "PE Purpose",
    "PE Purpose - Sub2",
    "Activity Initiator",
    "Activity Owner",
    "Activity Executor",
    "Deploy Approval",
    "NPO KPI Approval",
    "NPO AT Approval",
    "TSS Approval",
    "Status",
    "Domain",
    "Remarks1",
    "Remarks2",
    "Remarks",
    "Created At (UTC)",
]

# -----------------------------
# Flask app
# -----------------------------
app = Flask(__name__, template_folder="templates", static_folder="static")

# -----------------------------
# Azure / Graph Config
# IMPORTANT:
# Put these in environment variables, NOT hard-coded in source code
# -----------------------------
CLIENT_ID = os.getenv("CLIENT_ID", "e79ecadf-7da2-4ff3-9ca6-f3738141952e").strip()
TENANT_ID = os.getenv("TENANT_ID", "5d471751-9675-428d-917b-70f44f9630b0").strip()
CLIENT_SECRET = os.getenv("CLIENT_SECRET", "PT58Q~dEf151~EXjh9L2HxVYNZKJZqWNTHoWMaE2").strip()

# Example: user@yourcompany.com OR Azure AD user object id
ONEDRIVE_USER = os.getenv("ONEDRIVE_USER", "").strip()

# Folder inside that user's OneDrive
ONEDRIVE_FOLDER = os.getenv("ONEDRIVE_FOLDER", "Rudra/PEData").strip().strip("/")

AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
SCOPES = ["https://graph.microsoft.com/.default"]


# -----------------------------
# Helpers
# -----------------------------
def validate_config() -> None:
    missing = []
    if not CLIENT_ID:
        missing.append("CLIENT_ID")
    if not TENANT_ID:
        missing.append("TENANT_ID")
    if not CLIENT_SECRET:
        missing.append("CLIENT_SECRET")
    if not ONEDRIVE_USER:
        missing.append("ONEDRIVE_USER")

    if missing:
        raise RuntimeError(
            "Missing required environment variables: " + ", ".join(missing)
        )


def ensure_workbook() -> None:
    EXCEL_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)

    if not EXCEL_FILE_PATH.exists():
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "PE Data"
        worksheet.append(HEADERS)
        workbook.save(EXCEL_FILE_PATH)
        workbook.close()


def list_to_string(value):
    if isinstance(value, list):
        return ", ".join(str(item).strip() for item in value if str(item).strip())
    if value is None:
        return ""
    return str(value).strip()


def get_next_sno() -> int:
    ensure_workbook()
    workbook = load_workbook(EXCEL_FILE_PATH)
    worksheet = workbook.active
    next_sno = max(worksheet.max_row, 1)
    workbook.close()
    return next_sno


def get_graph_token() -> str:
    """
    Acquire app-only token using client credentials flow.
    """
    validate_config()

    app_msal = msal.ConfidentialClientApplication(
        CLIENT_ID,
        authority=AUTHORITY,
        client_credential=CLIENT_SECRET,
    )

    result = app_msal.acquire_token_for_client(scopes=SCOPES)

    if "access_token" not in result:
        raise RuntimeError(
            "OneDrive token error: "
            f"error={result.get('error')} | "
            f"description={result.get('error_description')} | "
            f"correlation_id={result.get('correlation_id')}"
        )

    return result["access_token"]


def graph_request(method: str, url: str, token: str, **kwargs):
    headers = kwargs.pop("headers", {})
    headers.update({
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
    })

    response = requests.request(method, url, headers=headers, timeout=60, **kwargs)
    return response


def ensure_onedrive_folder(token: str, folder_path: str) -> None:
    """
    Ensures nested folder path exists in target user's OneDrive.
    Example folder_path: Rudra/PEData
    """
    if not folder_path:
        return

    parts = [p.strip() for p in folder_path.split("/") if p.strip()]
    current_path = ""

    for part in parts:
        next_path = f"{current_path}/{part}" if current_path else part

        check_url = (
            f"https://graph.microsoft.com/v1.0/users/{quote(ONEDRIVE_USER, safe='')}"
            f"/drive/root:/{quote(next_path, safe='/')}"
        )
        check_resp = graph_request("GET", check_url, token)

        if check_resp.status_code == 200:
            current_path = next_path
            continue

        if check_resp.status_code != 404:
            try:
                err_json = check_resp.json()
            except Exception:
                err_json = check_resp.text
            raise RuntimeError(f"Folder check failed for '{next_path}': {err_json}")

        # Create missing folder
        if current_path:
            create_url = (
                f"https://graph.microsoft.com/v1.0/users/{quote(ONEDRIVE_USER, safe='')}"
                f"/drive/root:/{quote(current_path, safe='/')}:/children"
            )
        else:
            create_url = (
                f"https://graph.microsoft.com/v1.0/users/{quote(ONEDRIVE_USER, safe='')}"
                f"/drive/root/children"
            )

        create_body = {
            "name": part,
            "folder": {},
            "@microsoft.graph.conflictBehavior": "replace"
        }

        create_resp = graph_request("POST", create_url, token, json=create_body)

        if create_resp.status_code not in (200, 201):
            try:
                err_json = create_resp.json()
            except Exception:
                err_json = create_resp.text
            raise RuntimeError(f"Folder create failed for '{next_path}': {err_json}")

        current_path = next_path


def upload_to_onedrive() -> dict:
    """
    Upload local Excel file to target user's OneDrive folder.
    Returns uploaded item details.
    """
    token = get_graph_token()

    # Ensure target folder exists
    ensure_onedrive_folder(token, ONEDRIVE_FOLDER)

    relative_path = f"{ONEDRIVE_FOLDER}/{EXCEL_FILE_PATH.name}" if ONEDRIVE_FOLDER else EXCEL_FILE_PATH.name

    upload_url = (
        f"https://graph.microsoft.com/v1.0/users/{quote(ONEDRIVE_USER, safe='')}"
        f"/drive/root:/{quote(relative_path, safe='/')}:/content"
    )

    with open(EXCEL_FILE_PATH, "rb") as f:
        response = graph_request(
            "PUT",
            upload_url,
            token,
            headers={"Content-Type": "application/octet-stream"},
            data=f,
        )

    if response.status_code not in (200, 201):
        try:
            err_json = response.json()
        except Exception:
            err_json = response.text
        raise RuntimeError(f"OneDrive upload failed: {err_json}")

    item = response.json()

    return {
        "id": item.get("id"),
        "name": item.get("name"),
        "web_url": item.get("webUrl"),
        "size": item.get("size"),
    }


def save_row_to_excel(data: dict) -> None:
    ensure_workbook()

    workbook = load_workbook(EXCEL_FILE_PATH)
    worksheet = workbook.active

    row = [
        data.get("sno", ""),
        list_to_string(data.get("peNonPe")),
        list_to_string(data.get("category")),
        list_to_string(data.get("executor")),
        data.get("totalNodeCount", ""),
        list_to_string(data.get("siteCell")),
        list_to_string(data.get("circle")),
        data.get("dateOfPe", ""),
        list_to_string(data.get("pePurpose")),
        list_to_string(data.get("pePurposeSub2")),
        list_to_string(data.get("activityInitiator")),
        list_to_string(data.get("activityOwner")),
        list_to_string(data.get("activityExecutor")),
        list_to_string(data.get("deployApproval")),
        list_to_string(data.get("npoKpiApproval")),
        list_to_string(data.get("npoAtApproval")),
        list_to_string(data.get("tssApproval")),
        list_to_string(data.get("status")),
        list_to_string(data.get("domain")),
        list_to_string(data.get("remarks1")),
        list_to_string(data.get("remarks2")),
        data.get("remarks", ""),
        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
    ]

    worksheet.append(row)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def validate_payload(data: dict):
    required_single = {
        "peNonPe": "PE / Non-PE",
        "category": "Category",
        "executor": "Executor",
        "siteCell": "Site / Cell",
    }

    if not str(data.get("sno", "")).strip():
        return "S.No missing hai."

    for key, label in required_single.items():
        value = data.get(key, [])
        if isinstance(value, list):
            if not value:
                return f"{label} select karein."
        elif not str(value).strip():
            return f"{label} select karein."

    return None


# -----------------------------
# Routes
# -----------------------------
@app.get("/")
def home():
    ensure_workbook()
    return render_template("Form.html")


@app.get("/health")
def health():
    try:
        validate_config()
        ensure_workbook()
        return jsonify({
            "status": "ok",
            "local_excel": str(EXCEL_FILE_PATH),
            "onedrive_user": ONEDRIVE_USER,
            "onedrive_folder": ONEDRIVE_FOLDER,
        })
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.get("/next-sno")
def next_sno():
    try:
        return jsonify({"next_sno": get_next_sno()})
    except Exception as exc:
        print("ERROR in /next-sno:", exc)
        traceback.print_exc()
        return jsonify({"next_sno": "", "error": str(exc)}), 500


@app.post("/submit")
def submit():
    try:
        data = request.get_json(force=True) or {}

        error_message = validate_payload(data)
        if error_message:
            return jsonify({"success": False, "message": error_message}), 400

        # Step 1: Save locally
        save_row_to_excel(data)

        # Step 2: Upload to OneDrive
        try:
            uploaded = upload_to_onedrive()
            return jsonify({
                "success": True,
                "message": "Data local Excel me save ho gaya aur OneDrive pe upload bhi ho gaya.",
                "local_file": str(EXCEL_FILE_PATH),
                "onedrive_file": uploaded,
            })
        except Exception as upload_exc:
            print("ERROR in OneDrive upload:", upload_exc)
            traceback.print_exc()

            return jsonify({
                "success": False,
                "saved_local": True,
                "message": "Data local Excel me save ho gaya, lekin OneDrive upload fail ho gaya.",
                "local_file": str(EXCEL_FILE_PATH),
                "upload_error": str(upload_exc),
            }), 502

    except Exception as exc:
        print("ERROR in /submit:", exc)
        traceback.print_exc()
        return jsonify({"success": False, "message": str(exc)}), 500


if __name__ == "__main__":
    ensure_workbook()
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port, debug=False)
